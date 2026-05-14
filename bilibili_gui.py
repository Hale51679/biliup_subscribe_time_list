"""
B站关注列表导出工具 - GUI版本 v2
打包命令：
  pyinstaller --onefile --windowed --name "B站关注导出" --icon "bilibili_icon.ico" bilibili_gui.py
"""

# ── DPI感知（必须在任何tkinter导入之前）──
import ctypes, sys
try:
    if sys.platform == "win32":
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import requests
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageTk
import qrcode

# ── 配色方案 ──
PRIMARY    = "#00A1D6"
PRIMARY_DK = "#0085B3"
PRIMARY_LT = "#F0F9FF"
BG         = "#F4F5F7"
CARD       = "#FFFFFF"
TEXT       = "#1F2A3A"
TEXT_SUB   = "#8E9AAF"
BORDER     = "#E8ECF1"
DANGER     = "#F25D8E"
SUCCESS    = "#00A1D6"
WHITE      = "#FFFFFF"

# ───────────────────────────── 爬虫核心逻辑 ─────────────────────────────

def make_headers(sessdata):
    sessdata = sessdata.strip()
    if sessdata.lower().startswith("sessdata="):
        sessdata = sessdata[9:]
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.bilibili.com",
        "Cookie": f"SESSDATA={sessdata}",
    }

def get_following_list(uid, headers, log):
    followings = []
    page, page_size = 1, 50
    while True:
        url = f"https://api.bilibili.com/x/relation/followings?vmid={uid}&pn={page}&ps={page_size}&order=desc"
        resp = requests.get(url, headers=headers, timeout=10)
        data = resp.json()
        if data.get("code") != 0:
            log(f"❌ 获取关注列表失败: {data.get('message')}")
            return []
        items = data["data"].get("list", [])
        if not items:
            break
        for item in items:
            followings.append({"mid": item["mid"], "name": item["uname"]})
        total = data["data"].get("total", 0)
        log(f"📋 已获取 {len(followings)}/{total} 个UP主")
        if len(followings) >= total:
            break
        page += 1
        time.sleep(0.5)
    return followings

def get_follow_time(target_uid, headers, log, max_retries=5):
    url = f"https://api.bilibili.com/x/space/acc/relation?mid={target_uid}"
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            data = resp.json()
            code = data.get("code")
            if code == -799:
                wait = attempt * 2
                log(f"  ⚠️ 风控限流，等待 {wait}s 后重试（{attempt}/{max_retries}）...")
                time.sleep(wait)
                continue
            if code == 0:
                relation = data.get("data", {}).get("relation")
                if not relation:
                    return None, False
                mtime = relation.get("mtime", 0)
                attribute = relation.get("attribute", 0)
                is_following = attribute != 0
                return (mtime if mtime and mtime > 0 else None), is_following
            log(f"  ⚠️ uid={target_uid} 异常 code={code}: {data.get('message')}")
            return None, False
        except Exception as e:
            log(f"  ⚠️ 查询 {target_uid} 出错: {e}")
            time.sleep(2)
    log(f"  ❌ uid={target_uid} 重试 {max_retries} 次后仍失败")
    return None, False

def get_user_info(uid, headers):
    url = f"https://api.bilibili.com/x/space/acc/info?mid={uid}"
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        data = resp.json()
        if data.get("code") == 0:
            return data["data"].get("name", "未知用户")
    except Exception:
        pass
    return "未知用户"

def timestamp_to_str(ts):
    if ts:
        return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    return "未知"

def export_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "关注列表"
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="00A1D6")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["序号", "UP主名称", "UID", "关注时间"]
    col_widths = [8, 30, 18, 22]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28
    alt_fill = PatternFill("solid", start_color="F0F9FF")
    for i, row in enumerate(data, 1):
        excel_row = i + 1
        values = [i, row["name"], row["mid"], row["follow_time"]]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = cell_align if col != 1 else Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[excel_row].height = 20
    ws.freeze_panes = "A2"
    wb.save(filename)

# ───────────────────────────── 扫码登录 ─────────────────────────────

QR_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.bilibili.com/",
    "Origin": "https://www.bilibili.com",
}

def qr_generate():
    resp = requests.get(
        "https://passport.bilibili.com/x/passport-login/web/qrcode/generate?source=main-fe-header",
        headers=QR_HEADERS, timeout=10
    )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"生成二维码失败: {data.get('message')}")
    return data["data"]["qrcode_key"], data["data"]["url"]

def qr_poll(qrcode_key):
    resp = requests.get(
        f"https://passport.bilibili.com/x/passport-login/web/qrcode/poll?qrcode_key={qrcode_key}",
        headers=QR_HEADERS, timeout=10
    )
    data = resp.json()
    # 顶层 code 始终为 0，实际状态在 data.data.code 中
    inner_code = data.get("data", {}).get("code", -1)
    if inner_code == 0:
        # 登录成功！从 cookie 中提取 SESSDATA
        sessdata = resp.cookies.get("SESSDATA", "")
        if not sessdata:
            # 兜底：手动从 Set-Cookie 解析
            for cookie in resp.raw.headers.getlist("Set-Cookie"):
                if cookie.strip().startswith("SESSDATA="):
                    sessdata = cookie.split("SESSDATA=")[1].split(";")[0]
                    break
        return {"status": "ok", "sessdata": sessdata, "data": data}
    elif inner_code == 86038:
        return {"status": "expired"}
    elif inner_code == 86090:
        return {"status": "scanned"}
    else:
        return {"status": "waiting"}

def get_my_uid(sessdata):
    """用 SESSDATA 查询当前登录用户的 UID"""
    headers = {
        **QR_HEADERS,
        "Cookie": f"SESSDATA={sessdata}",
    }
    resp = requests.get(
        "https://api.bilibili.com/x/web-interface/nav",
        headers=headers, timeout=10
    )
    data = resp.json()
    if data.get("code") == 0:
        return data["data"].get("mid")
    return None

# ───────────────────────────── GUI ─────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("B站关注列表导出工具")
        self.resizable(True, True)
        self.minsize(520, 640)
        # 设置窗口图标（打包后从exe同目录读取，开发时从脚本同目录读取）
        try:
            import os
            base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base, "bilibili_icon.ico")
            self.iconbitmap(icon_path)
        except Exception:
            pass
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_reqwidth(), self.winfo_reqheight()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _entry(self, parent, textvariable=None, show=None):
        kw = dict(font=("Consolas", 10), relief="flat", bg=WHITE,
                  highlightthickness=1, highlightbackground=BORDER, highlightcolor=PRIMARY,
                  insertbackground=TEXT, selectbackground=PRIMARY_LT, selectforeground=TEXT)
        if textvariable: kw["textvariable"] = textvariable
        if show: kw["show"] = show
        return tk.Entry(parent, **kw)

    def _text_box(self, parent, height=3):
        return tk.Text(parent, height=height, font=("Consolas", 10), relief="flat",
                       bg=WHITE, highlightthickness=1, highlightbackground=BORDER,
                       highlightcolor=PRIMARY, wrap="word", insertbackground=TEXT)

    def _btn(self, parent, text, command, small=False, secondary=False):
        bg = WHITE if secondary else PRIMARY
        fg = PRIMARY if secondary else WHITE
        active_bg = BORDER if secondary else PRIMARY_DK
        active_fg = PRIMARY_DK if secondary else WHITE
        padx = 18 if small else 28
        pady = 4 if small else 6
        b = tk.Button(parent, text=text, padx=padx, pady=pady,
                      font=("微软雅黑", 9 if small else 11, "bold"),
                      bg=bg, fg=fg, relief="flat", cursor="hand2",
                      activebackground=active_bg, activeforeground=active_fg,
                      highlightthickness=1, highlightbackground=bg, bd=0, command=command)
        b.bind("<Enter>", lambda e: b.configure(bg=PRIMARY_LT if secondary else PRIMARY_DK))
        b.bind("<Leave>", lambda e: b.configure(bg=bg))
        return b

    def _card(self, parent, **kw):
        return tk.Frame(parent, bg=CARD, highlightthickness=1, highlightbackground=BORDER, **kw)

    def _log_widget(self, parent, height=6):
        frame = tk.Frame(parent, bg=BG)
        box = tk.Text(frame, height=height, font=("Consolas", 9), bg="#1A1D23", fg="#CCCCCC",
                      relief="flat", state="disabled", wrap="word",
                      insertbackground="#CCCCCC", padx=8, pady=6)
        scroll = tk.Scrollbar(frame, command=box.yview, bg="#2A2D33", troughcolor="#1A1D23",
                              activebackground="#555", relief="flat", bd=0)
        box.configure(yscrollcommand=scroll.set)
        box.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        return frame, box

    def _section_label(self, parent, text):
        return tk.Label(parent, text=text, font=("微软雅黑", 10, "bold"), bg=BG, fg=TEXT, anchor="w")

    def _build_ui(self):
        self.configure(bg=BG)

        # 标题头
        title_frame = tk.Frame(self, bg=PRIMARY, pady=16)
        title_frame.pack(fill="x")
        tk.Label(title_frame, text="🎬 B站关注列表导出工具",
                 font=("微软雅黑", 17, "bold"), bg=PRIMARY, fg=WHITE).pack()
        tk.Label(title_frame, text="批量导出你关注的UP主名称、UID与关注时间",
                 font=("微软雅黑", 9), bg=PRIMARY, fg="#CCE8F7").pack()

        # Tab样式
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", font=("微软雅黑", 10),
                        padding=[18, 7], background="#DEE2E6", foreground=TEXT_SUB)
        style.map("TNotebook.Tab",
                  background=[("selected", PRIMARY)],
                  foreground=[("selected", WHITE)],
                  lightcolor=[("selected", PRIMARY)])
        style.layout("TNotebook.Tab", [
            ("Notebook.tab", {"sticky": "nswe", "children":
                [("Notebook.padding", {"side": "top", "sticky": "nswe", "children":
                    [("Notebook.label", {"side": "top", "sticky": ""})]
                })]
            })
        ])

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=0)
        self.nb = nb

        tab_qr     = tk.Frame(nb, bg=BG)
        tab_batch  = tk.Frame(nb, bg=BG)
        tab_single = tk.Frame(nb, bg=BG)
        nb.add(tab_qr,     text="  扫码登录  ")
        nb.add(tab_batch,  text="  批量导出  ")
        nb.add(tab_single, text="  单个查询  ")

        self._build_qr_tab(tab_qr)
        self._build_batch_tab(tab_batch)
        self._build_single_tab(tab_single)

    # ── 扫码登录 ────────────────────────────────────────────────────

    def _build_qr_tab(self, parent):
        p = dict(padx=24)
        container = tk.Frame(parent, bg=BG)
        container.pack(fill="both", expand=True, **p)
        tk.Frame(container, bg=BG, height=6).pack()

        # 二维码显示区
        qr_card = self._card(container, width=220, height=220)
        qr_card.pack(pady=(0, 10))
        qr_card.pack_propagate(False)
        self.qr_label = tk.Label(qr_card, bg=CARD)
        self.qr_label.pack(expand=True)
        self.qr_placeholder = tk.Label(qr_card, text="点击下方按钮\n生成二维码",
                                       font=("微软雅黑", 10), bg=CARD, fg=TEXT_SUB)
        self.qr_placeholder.place(relx=0.5, rely=0.5, anchor="center")

        # 状态提示
        self.qr_status_var = tk.StringVar(value="就绪")
        tk.Label(container, textvariable=self.qr_status_var,
                 font=("微软雅黑", 9), bg=BG, fg=TEXT_SUB).pack()

        # 按钮行
        btn_frame = tk.Frame(container, bg=BG)
        btn_frame.pack(pady=(8, 6))
        self.qr_gen_btn = self._btn(btn_frame, "  生成二维码  ", self._start_qr)
        self.qr_gen_btn.pack()

        # 结果卡片
        card = self._card(container)
        card.pack(fill="x")
        inner = tk.Frame(card, bg=CARD, padx=16, pady=10)
        inner.pack(fill="x")

        tk.Label(inner, text="SESSDATA", font=("微软雅黑", 8), bg=CARD, fg=TEXT_SUB).pack(anchor="w")
        self.qr_sessdata_var = tk.StringVar(value="")
        tk.Label(inner, textvariable=self.qr_sessdata_var,
                 font=("Consolas", 10), bg=CARD, fg=TEXT,
                 wraplength=440, justify="left").pack(anchor="w", fill="x")

        tk.Frame(inner, bg=BORDER, height=1).pack(fill="x", pady=6)

        tk.Label(inner, text="扫码后自动填入以下位置：", font=("微软雅黑", 8), bg=CARD, fg=TEXT_SUB).pack(anchor="w")
        self.qr_targets_var = tk.StringVar(value="—")
        tk.Label(inner, textvariable=self.qr_targets_var,
                 font=("微软雅黑", 9), bg=CARD, fg=PRIMARY).pack(anchor="w")

        log_frame, self.qr_log_box = self._log_widget(container, height=5)
        log_frame.pack(fill="both", expand=True, pady=(10, 0))

        self.qr_polling = False
        self.qr_tk_image = None

    def _build_batch_tab(self, parent):
        p = dict(padx=24)
        container = tk.Frame(parent, bg=BG)
        container.pack(fill="both", expand=True, **p)
        tk.Frame(container, bg=BG, height=4).pack()

        # UID
        self._section_label(container, "你的B站 UID").pack(fill="x")
        self.uid_var = tk.StringVar()
        self._entry(container, textvariable=self.uid_var).pack(fill="x", ipady=5, pady=(4, 12))

        # SESSDATA
        self._section_label(container, "SESSDATA").pack(fill="x")
        self.sessdata_text = self._text_box(container, height=3)
        self.sessdata_text.pack(fill="x", pady=(4, 12))

        # 导出路径
        self._section_label(container, "导出路径").pack(fill="x")
        pf = tk.Frame(container, bg=BG)
        pf.pack(fill="x", pady=(4, 10))
        self.path_var = tk.StringVar(value="b站关注列表.xlsx")
        self._entry(pf, textvariable=self.path_var).pack(side="left", fill="x", expand=True, ipady=5)
        self._btn(pf, "浏览", self._browse, small=True, secondary=True).pack(side="left", padx=(6, 0))

        # 进度
        self.b_progress = ttk.Progressbar(container, mode="determinate")
        self.b_progress.pack(fill="x")
        self.b_status_var = tk.StringVar(value="就绪")
        tk.Label(container, textvariable=self.b_status_var,
                 font=("微软雅黑", 9), bg=BG, fg=TEXT_SUB).pack(anchor="w", pady=(2, 0))

        log_frame, self.b_log_box = self._log_widget(container, height=5)
        log_frame.pack(fill="both", expand=True, pady=(8, 0))

        self.run_btn = self._btn(container, "  开始导出  ", self._start_batch)
        self.run_btn.pack(pady=12)

    # ── 单个查询 ──────────────────────────────────────────────────────

    def _build_single_tab(self, parent):
        p = dict(padx=24)
        container = tk.Frame(parent, bg=BG)
        container.pack(fill="both", expand=True, **p)
        tk.Frame(container, bg=BG, height=4).pack()

        # SESSDATA
        self._section_label(container, "SESSDATA").pack(fill="x")
        self.s_sessdata_text = self._text_box(container, height=3)
        self.s_sessdata_text.pack(fill="x", pady=(4, 10))

        # UP主UID + 查询按钮
        qf = tk.Frame(container, bg=BG)
        qf.pack(fill="x", pady=(0, 10))
        tk.Label(qf, text="UP主 UID", font=("微软雅黑", 10, "bold"),
                 bg=BG, fg=TEXT).pack(anchor="w")
        ef = tk.Frame(qf, bg=BG)
        ef.pack(fill="x", pady=(4, 0))
        self.query_var = tk.StringVar()
        qe = self._entry(ef, textvariable=self.query_var)
        qe.pack(side="left", fill="x", expand=True, ipady=5)
        qe.bind("<Return>", lambda e: self._start_single())
        self._btn(ef, "查询", self._start_single, small=True).pack(side="left", padx=(6, 0))

        # 结果卡片
        card = self._card(container)
        card.pack(fill="x")
        inner = tk.Frame(card, bg=CARD, padx=16, pady=14)
        inner.pack(fill="x")

        def card_col(row_frame, label, var, col):
            f = tk.Frame(row_frame, bg=CARD)
            f.grid(row=0, column=col, sticky="w", padx=(0, 36))
            tk.Label(f, text=label, font=("微软雅黑", 8), bg=CARD, fg=TEXT_SUB).pack(anchor="w")
            lbl = tk.Label(f, textvariable=var, font=("微软雅黑", 13, "bold"), bg=CARD, fg=TEXT)
            lbl.pack(anchor="w")
            row_frame.columnconfigure(col, weight=1)
            return lbl

        row1 = tk.Frame(inner, bg=CARD); row1.pack(fill="x")
        self.s_name_var = tk.StringVar(value="—")
        self.s_uid_var  = tk.StringVar(value="—")
        card_col(row1, "UP主名称", self.s_name_var, 0)
        card_col(row1, "UID",     self.s_uid_var,  1)

        tk.Frame(inner, bg=BORDER, height=1).pack(fill="x", pady=8)

        row2 = tk.Frame(inner, bg=CARD); row2.pack(fill="x")
        self.s_status_var = tk.StringVar(value="—")
        self.s_time_var   = tk.StringVar(value="—")
        self.s_status_lbl = card_col(row2, "关注状态", self.s_status_var, 0)
        card_col(row2, "关注时间", self.s_time_var, 1)

        log_frame, self.s_log_box = self._log_widget(container, height=6)
        log_frame.pack(fill="both", expand=True, pady=(10, 0))

    # ── 通用工具 ──────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="b站关注列表.xlsx", title="选择导出位置")
        if path:
            self.path_var.set(path)

    def _log(self, box, msg):
        box.configure(state="normal")
        box.insert("end", msg + "\n")
        box.see("end")
        box.configure(state="disabled")

    def _clear_log(self, box):
        box.configure(state="normal")
        box.delete("1.0", "end")
        box.configure(state="disabled")

    # ── 扫码登录线程 ────────────────────────────────────────────────

    def _start_qr(self):
        self.qr_gen_btn.configure(state="disabled", text="  生成中...  ")
        self.qr_status_var.set("正在请求二维码...")
        self._clear_log(self.qr_log_box)
        self.qr_sessdata_var.set("")
        self.qr_targets_var.set("—")
        # 清除旧二维码
        self.qr_label.config(image="")
        self.qr_label.image = None
        self.qr_tk_image = None
        self.qr_placeholder.place(relx=0.5, rely=0.5, anchor="center")
        threading.Thread(target=self._run_qr_generate, daemon=True).start()

    def _run_qr_generate(self):
        log = lambda m: self._log(self.qr_log_box, m)
        try:
            log("📱 正在向B站请求二维码...")
            qrcode_key, url = qr_generate()
            log(f"✅ 二维码已生成")
            log(f"🔑 qrcode_key: {qrcode_key}")
            log("💡 请打开B站App扫描二维码")
            self.qr_status_var.set("请打开B站App扫码")
            self.after(0, self._show_qr_image, url)
            self.qr_polling = True
            self.qr_gen_btn.configure(text="  重新生成  ", state="normal")
            self._run_qr_poll(qrcode_key, log)
        except Exception as e:
            log(f"❌ 生成二维码失败: {e}")
            self.qr_status_var.set("生成失败")
            self.qr_gen_btn.configure(state="normal", text="  重新生成  ")

    def _show_qr_image(self, url):
        # 移除占位文字
        self.qr_placeholder.place_forget()
        # 生成二维码 PIL Image
        pil_img = qrcode.make(url, box_size=6).convert("RGB")
        # 加白边
        w, h = pil_img.size
        padded = Image.new("RGB", (w + 20, h + 20), "white")
        padded.paste(pil_img, (10, 10, w + 10, h + 10))
        # 缩放到 200x200 显示
        padded = padded.resize((200, 200), Image.NEAREST)
        self.qr_tk_image = ImageTk.PhotoImage(padded)
        self.qr_label.config(image=self.qr_tk_image)

    def _run_qr_poll(self, qrcode_key, log):
        timeout = 180
        start = time.time()
        while self.qr_polling and (time.time() - start) < timeout:
            try:
                result = qr_poll(qrcode_key)
                if result["status"] == "ok":
                    sessdata = result["sessdata"]
                    log(f"\n🎉 扫码登录成功！")
                    # 获取 UID
                    uid = get_my_uid(sessdata)
                    if uid:
                        log(f"👤 UID: {uid}")
                    else:
                        log("⚠️ 未能获取 UID，请手动填写")
                    # 自动填入其他Tab
                    self.after(0, self._fill_login_info, sessdata, uid)
                    self.qr_polling = False
                    return
                elif result["status"] == "scanned":
                    self.qr_status_var.set("已扫码，请在手机上确认...")
                    log("📱 已扫码，等待确认...")
                elif result["status"] == "expired":
                    log("⏰ 二维码已过期，请点击「重新生成」")
                    self.qr_status_var.set("二维码已过期")
                    self.qr_polling = False
                    return
                # else: waiting, continue polling
                time.sleep(1.5)
            except Exception as e:
                log(f"⚠️ 轮询异常: {e}")
                time.sleep(2)

        if self.qr_polling:
            log("⏰ 扫码超时，请重新生成二维码")
            self.qr_status_var.set("扫码超时")
            self.qr_polling = False

    def _fill_login_info(self, sessdata, uid):
        """将 SESSDATA 和 UID 自动填入其他Tab"""
        # SESSDATA → 批量导出
        self.sessdata_text.delete("1.0", "end")
        self.sessdata_text.insert("1.0", sessdata)
        # SESSDATA → 单个查询
        self.s_sessdata_text.delete("1.0", "end")
        self.s_sessdata_text.insert("1.0", sessdata)
        # UID → 批量导出
        if uid:
            self.uid_var.set(str(uid))
        self.qr_sessdata_var.set(sessdata)
        self.qr_status_var.set("✅ 登录成功！")
        parts = ["已填入「批量导出」的 SESSDATA"]
        if uid:
            parts.append("和 UID")
        parts.append("，以及「单个查询」的 SESSDATA")
        self.qr_targets_var.set("✅ " + "".join(parts))
        # 2秒后自动切换到批量导出
        self.after(2000, self._switch_to_batch)

    def _switch_to_batch(self):
        self.nb.select(1)

    def _stop_qr(self):
        self.qr_polling = False

    # ── 批量导出线程 ──────────────────────────────────────────────────

    def _start_batch(self):
        uid = self.uid_var.get().strip()
        sessdata = self.sessdata_text.get("1.0", "end").strip()
        path = self.path_var.get().strip()
        if not uid or not sessdata:
            messagebox.showerror("缺少信息", "请填写 UID 和 SESSDATA")
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"
            self.path_var.set(path)
        self.run_btn.configure(state="disabled", text="  导出中...  ")
        self.b_progress["value"] = 0
        self._clear_log(self.b_log_box)
        threading.Thread(target=self._run_batch, args=(uid, sessdata, path), daemon=True).start()

    def _run_batch(self, uid, sessdata, path):
        headers = make_headers(sessdata)
        log = lambda m: self._log(self.b_log_box, m)
        try:
            self.b_status_var.set("正在获取关注列表...")
            followings = get_following_list(uid, headers, log)
            if not followings:
                messagebox.showerror("失败", "未能获取关注列表，请检查 UID 和 SESSDATA 是否正确")
                return
            total = len(followings)
            log(f"\n✅ 共获取 {total} 个UP主，开始查询关注时间...\n")
            results = []
            for i, up in enumerate(followings, 1):
                mtime, _ = get_follow_time(up["mid"], headers, log)
                follow_time = timestamp_to_str(mtime)
                results.append({"name": up["name"], "mid": up["mid"], "follow_time": follow_time})
                log(f"  [{i:3d}/{total}] {up['name'][:16]:<16} → {follow_time}")
                self.b_status_var.set(f"查询关注时间 {i}/{total}")
                self.b_progress["value"] = i / total * 100
                time.sleep(0.8)
            self.b_status_var.set("正在写入 Excel...")
            export_to_excel(results, path)
            log(f"\n🎉 导出完成！共 {len(results)} 条记录")
            log(f"📁 文件已保存到：{path}")
            self.b_status_var.set(f"完成！已导出 {len(results)} 条记录")
            messagebox.showinfo("完成", f"导出成功！共 {len(results)} 条记录\n\n文件位置：\n{path}")
        except Exception as e:
            log(f"\n❌ 发生错误: {e}")
            messagebox.showerror("错误", str(e))
        finally:
            self.run_btn.configure(state="normal", text="  开始导出  ")

    # ── 单个查询线程 ──────────────────────────────────────────────────

    def _start_single(self):
        target_uid = self.query_var.get().strip()
        sessdata = self.s_sessdata_text.get("1.0", "end").strip()
        if not target_uid or not sessdata:
            messagebox.showerror("缺少信息", "请填写 SESSDATA 和要查询的UP主UID")
            return
        self._clear_log(self.s_log_box)
        self.s_name_var.set("查询中...")
        self.s_uid_var.set(target_uid)
        self.s_status_var.set("—")
        self.s_time_var.set("—")
        threading.Thread(target=self._run_single, args=(target_uid, sessdata), daemon=True).start()

    def _run_single(self, target_uid, sessdata):
        headers = make_headers(sessdata)
        log = lambda m: self._log(self.s_log_box, m)
        try:
            log(f"🔍 正在查询 uid={target_uid}...")
            name = get_user_info(target_uid, headers)
            mtime, is_following = get_follow_time(target_uid, headers, log)
            self.s_name_var.set(name)
            self.s_uid_var.set(str(target_uid))
            if not is_following:
                self.s_status_var.set("未关注")
                self.s_status_lbl.configure(fg="#f25d8e")
                self.s_time_var.set("—")
                log(f"✅ {name}（{target_uid}）：你尚未关注该UP主")
            else:
                self.s_status_var.set("已关注 ✓")
                self.s_status_lbl.configure(fg="#00a1d6")
                time_str = timestamp_to_str(mtime)
                self.s_time_var.set(time_str)
                log(f"✅ {name}（{target_uid}）：关注于 {time_str}")
        except Exception as e:
            log(f"❌ 查询失败: {e}")
            self.s_name_var.set("查询失败")


if __name__ == "__main__":
    app = App()
    app.mainloop()
