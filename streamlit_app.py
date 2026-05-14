"""
B站关注列表导出工具 - Web版
部署到 Streamlit Community Cloud：绑定本仓库，启动入口 streamlit_app.py
"""
import streamlit as st
import requests
import time
import datetime
import io
import urllib.parse
import qrcode
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 页面配置 ──────────────────────────────────────────────────────────
st.set_page_config(page_title="B站关注导出", page_icon="🎬", layout="centered")

# ── 样式 ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .stButton button { font-weight: 600; }
    .login-card {
        background: #f8f9fa; border-radius: 12px;
        padding: 1.5rem; margin: 1rem 0;
        border: 1px solid #e9ecef;
    }
    .success-card {
        background: #d4edda; border-radius: 8px;
        padding: 1rem; margin: 0.5rem 0;
        border: 1px solid #c3e6cb;
    }
    .sessdata-display {
        background: #f0f8ff; border-radius: 8px;
        padding: 0.75rem 1rem;
        border: 1px solid #b8daff;
        font-family: 'Courier New', monospace;
        font-size: 0.85rem;
        word-break: break-all;
    }
</style>
""", unsafe_allow_html=True)

# ── Session State 初始化 ──────────────────────────────────────────────
for key in ("sessdata", "uid", "qrcode_key", "qr_url", "qr_generated"):
    if key not in st.session_state:
        if key in ("qr_generated",):
            st.session_state[key] = False
        else:
            st.session_state[key] = "" if key in ("sessdata", "uid") else None

# ── 核心逻辑 ──────────────────────────────────────────────────────────

API_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www.bilibili.com/",
    "Origin": "https://www.bilibili.com",
}

def make_headers(sessdata):
    sessdata = sessdata.strip()
    if sessdata.lower().startswith("sessdata="):
        sessdata = sessdata[9:]
    return {
        "User-Agent": API_HEADERS["User-Agent"],
        "Referer": "https://www.bilibili.com",
        "Cookie": f"SESSDATA={sessdata}",
    }

def get_following_list(uid, headers, log_fn, max_retries=5):
    followings = []
    page, page_size = 1, 50
    while True:
        for attempt in range(1, max_retries + 1):
            try:
                url = f"https://api.bilibili.com/x/relation/followings?vmid={uid}&pn={page}&ps={page_size}&order=desc"
                resp = requests.get(url, headers=headers, timeout=10)
                data = resp.json()
                code = data.get("code")
                if code == -799:
                    wait = attempt * 3
                    log_fn(f"风控限流，等待 {wait}s（第{page}页，尝试 {attempt}/{max_retries}）...")
                    time.sleep(wait)
                    continue
                if code != 0:
                    log_fn(f"获取关注列表失败: {data.get('message')}")
                    return []
                items = data["data"].get("list", [])
                if not items:
                    return followings
                for item in items:
                    followings.append({"mid": item["mid"], "name": item["uname"]})
                total = data["data"].get("total", 0)
                log_fn(f"已获取 {len(followings)}/{total} 个UP主")
                break
            except Exception as e:
                log_fn(f"第{page}页出错: {e}")
                time.sleep(2)
        else:
            log_fn(f"第{page}页重试 {max_retries} 次后放弃")
            return followings
        if len(followings) >= total:
            break
        page += 1
        time.sleep(1)
    return followings

def get_follow_time(target_uid, headers, log_fn, max_retries=8):
    url = f"https://api.bilibili.com/x/space/acc/relation?mid={target_uid}"
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            data = resp.json()
            code = data.get("code")
            if code == -799:
                wait = min(attempt * 3, 30)
                log_fn(f"风控限流，等待 {wait}s（{attempt}/{max_retries}）...")
                time.sleep(wait)
                continue
            if code == 0:
                relation = data.get("data", {}).get("relation")
                if not relation:
                    return None, False
                mtime = relation.get("mtime", 0)
                attribute = relation.get("attribute", 0)
                return (mtime if mtime and mtime > 0 else None), attribute != 0
            log_fn(f"uid={target_uid} 异常 code={code}")
            return None, False
        except Exception as e:
            log_fn(f"查询 {target_uid} 出错: {e}")
            time.sleep(2)
    log_fn(f"uid={target_uid} 重试 {max_retries} 次后失败")
    return None, False

def get_user_info(uid, headers, max_retries=3):
    url = f"https://api.bilibili.com/x/space/acc/info?mid={uid}"
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            data = resp.json()
            code = data.get("code")
            if code == -799:
                time.sleep(attempt * 2)
                continue
            if code == 0:
                return data["data"].get("name", "未知用户")
            return "未知用户"
        except Exception:
            time.sleep(2)
    return "未知用户"

def timestamp_to_str(ts):
    if ts:
        return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    return "未知"

def export_to_excel_bytes(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "关注列表"
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="00A1D6")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["序号", "UP主名称", "UID", "关注时间"]
    col_widths = [8, 30, 18, 22]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28
    alt_fill = PatternFill("solid", start_color="F0F9FF")
    for i, row in enumerate(results, 1):
        excel_row = i + 1
        values = [i, row["name"], row["mid"], row["follow_time"]]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = cell_align if col != 1 else Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[excel_row].height = 20
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── 扫码逻辑 ──────────────────────────────────────────────────────────

def qr_generate():
    resp = requests.get(
        "https://passport.bilibili.com/x/passport-login/web/qrcode/generate?source=main-fe-header",
        headers=API_HEADERS, timeout=10
    )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"生成二维码失败: {data.get('message')}")
    return data["data"]["qrcode_key"], data["data"]["url"]

def qr_poll(qrcode_key):
    resp = requests.get(
        f"https://passport.bilibili.com/x/passport-login/web/qrcode/poll?qrcode_key={qrcode_key}",
        headers=API_HEADERS, timeout=10
    )
    data = resp.json()
    inner_code = data.get("data", {}).get("code", -1)
    if inner_code == 0:
        sessdata = resp.cookies.get("SESSDATA", "")
        if not sessdata:
            for cookie in resp.raw.headers.getlist("Set-Cookie"):
                if cookie.strip().startswith("SESSDATA="):
                    sessdata = cookie.split("SESSDATA=")[1].split(";")[0]
                    break
        sessdata = urllib.parse.unquote(sessdata) if sessdata else ""
        return {"status": "ok", "sessdata": sessdata}
    elif inner_code == 86038:
        return {"status": "expired"}
    elif inner_code == 86090:
        return {"status": "scanned"}
    else:
        return {"status": "waiting"}

def get_my_uid(sessdata):
    headers = {**API_HEADERS, "Cookie": f"SESSDATA={sessdata}"}
    resp = requests.get("https://api.bilibili.com/x/web-interface/nav", headers=headers, timeout=10)
    data = resp.json()
    if data.get("code") == 0:
        return data["data"].get("mid")
    return None

# ── 标题 ──────────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center; padding: 0.5rem 0 1rem 0;">
    <h1 style="color:#00A1D6; font-size:1.8rem; margin:0;">🎬 B站关注列表导出</h1>
    <p style="color:#8E9AAF; font-size:0.9rem; margin:0.25rem 0 0 0;">
        批量导出UP主名称、UID与关注时间
    </p>
</div>
""", unsafe_allow_html=True)

# 已登录提示
if st.session_state.sessdata:
    st.markdown(
        f'<div class="success-card">✅ 已登录 · <b>UID:</b> {st.session_state.uid or "未知"} · '
        f'SESSDATA 已就绪</div>',
        unsafe_allow_html=True,
    )

# ── Tab 页面 ──────────────────────────────────────────────────────────
tab_qr, tab_batch, tab_single = st.tabs(["📱 扫码登录", "📋 批量导出", "🔍 单个查询"])

# ═══════════════════════════════ Tab 1: 扫码登录 ═══════════════════════

with tab_qr:
    st.markdown("##### 扫码登录获取 SESSDATA")

    if st.button("🔲 生成二维码", use_container_width=True, type="primary"):
        with st.spinner("正在生成二维码..."):
            try:
                key, url = qr_generate()
                st.session_state.qrcode_key = key
                st.session_state.qr_url = url
                st.session_state.qr_generated = True
            except Exception as e:
                st.error(f"生成失败: {e}")
                st.session_state.qr_generated = False

    if st.session_state.get("qr_generated") and st.session_state.qr_url:
        # 显示二维码
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            pil_img = qrcode.make(st.session_state.qr_url, box_size=6).convert("RGB")
            buf = io.BytesIO()
            pil_img.save(buf, format="PNG")
            st.image(buf.getvalue(), width=220, caption="请用B站App扫码")

        # 轮询扫码状态
        if st.button("🔄 检查扫码状态", use_container_width=True):
            result = qr_poll(st.session_state.qrcode_key)
            if result["status"] == "ok":
                sessdata = result["sessdata"]
                st.session_state.sessdata = sessdata
                st.session_state.qr_generated = False
                st.balloons()
                st.success("✅ 扫码登录成功！")

                # 获取 UID
                with st.spinner("正在获取用户信息..."):
                    uid = get_my_uid(sessdata)
                    if uid:
                        st.session_state.uid = str(uid)
                        st.info(f"👤 UID: {uid}")
            elif result["status"] == "scanned":
                st.info("📱 已扫码，请在手机上确认登录...")
            elif result["status"] == "expired":
                st.error("⏰ 二维码已过期，请重新生成")
                st.session_state.qr_generated = False
            else:
                st.info("⏳ 等待扫码...")

    # 显示 SESSDATA
    if st.session_state.sessdata:
        st.markdown("##### 当前登录信息")
        st.markdown(
            f'<div class="sessdata-display">{st.session_state.sessdata}</div>',
            unsafe_allow_html=True,
        )
        if st.session_state.uid:
            st.markdown(f'UID: **{st.session_state.uid}**')

    st.divider()
    st.markdown(
        "<p style='color:#8E9AAF; font-size:0.8rem; text-align:center;'>"
        "💡 扫码后 SESSDATA 和 UID 会自动填入其它 Tab</p>",
        unsafe_allow_html=True,
    )

# ═══════════════════════════════ Tab 2: 批量导出 ═══════════════════════

with tab_batch:
    st.markdown("##### 批量导出关注列表")

    col1, col2 = st.columns(2)
    with col1:
        uid_input = st.text_input("你的 B站 UID", value=st.session_state.uid,
                                  help="从扫码登录自动获取，或手动填写")
    with col2:
        sessdata_input = st.text_input("SESSDATA", value=st.session_state.sessdata,
                                       type="password",
                                       help="从扫码登录自动获取，或手动填写")

    if st.button("📥 开始导出", use_container_width=True, type="primary"):
        if not uid_input or not sessdata_input:
            st.error("请先填写 UID 和 SESSDATA")
        else:
            log_placeholder = st.empty()
            progress_bar = st.progress(0, text="初始化...")
            logs = []

            def log_fn(msg):
                logs.append(msg)
                log_placeholder.code("\n".join(logs[-10:]), language="")

            try:
                headers = make_headers(sessdata_input)
                log_fn("正在获取关注列表...")
                followings = get_following_list(uid_input, headers, log_fn)
                if not followings:
                    st.error("未能获取关注列表，请检查 UID 和 SESSDATA")
                    st.stop()

                total = len(followings)
                log_fn(f"共 {total} 个UP主，开始查询关注时间...")
                results = []
                for i, up in enumerate(followings):
                    mtime, _ = get_follow_time(up["mid"], headers, log_fn)
                    follow_time = timestamp_to_str(mtime)
                    results.append({"name": up["name"], "mid": up["mid"], "follow_time": follow_time})
                    progress_bar.progress((i + 1) / total, text=f"{i+1}/{total}")
                    time.sleep(1.5)

                log_placeholder.empty()
                progress_bar.empty()

                excel_buf = export_to_excel_bytes(results)
                st.success(f"🎉 导出完成！共 {len(results)} 条记录")
                st.download_button(
                    "📁 下载 Excel 文件",
                    data=excel_buf,
                    file_name="b站关注列表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

                # 预览前 5 条
                with st.expander("预览前5条"):
                    for r in results[:5]:
                        st.text(f"{r['name']} (UID:{r['mid']}) → {r['follow_time']}")

            except Exception as e:
                st.error(f"发生错误: {e}")

# ═══════════════════════════════ Tab 3: 单个查询 ═══════════════════════

with tab_single:
    st.markdown("##### 查询是否关注某个UP主")

    sessdata_single = st.text_input("SESSDATA", value=st.session_state.sessdata,
                                    type="password", key="single_sessdata",
                                    help="从扫码登录自动获取，或手动填写")
    target_uid = st.text_input("UP主 UID", placeholder="输入要查询的UP主UID")

    if st.button("🔍 查询", use_container_width=True, type="primary"):
        if not sessdata_single or not target_uid:
            st.error("请填写 SESSDATA 和 UID")
        else:
            with st.spinner("查询中..."):
                try:
                    headers = make_headers(sessdata_single)
                    name = get_user_info(target_uid, headers)
                    mtime, is_following = get_follow_time(target_uid, headers, lambda m: None)

                    # 结果卡片
                    if is_following:
                        time_str = timestamp_to_str(mtime)
                        st.markdown(f"""
                        <div style="background:#e8f7fd; border-radius:10px; padding:1.2rem; border:1px solid #b8daff;">
                            <h4 style="margin:0 0 0.5rem 0; color:#00A1D6;">{name}</h4>
                            <p style="margin:0 0 0.25rem 0;"><b>UID:</b> {target_uid}</p>
                            <p style="margin:0; color:#28a745;">✅ 已关注 · {time_str}</p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div style="background:#fff3f3; border-radius:10px; padding:1.2rem; border:1px solid #f5c6cb;">
                            <h4 style="margin:0 0 0.5rem 0; color:#333;">{name}</h4>
                            <p style="margin:0 0 0.25rem 0;"><b>UID:</b> {target_uid}</p>
                            <p style="margin:0; color:#dc3545;">❌ 未关注</p>
                        </div>
                        """, unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"查询失败: {e}")

st.markdown("""
<div style="text-align:center; padding:1.5rem 0 0 0;">
    <p style="color:#8E9AAF; font-size:0.75rem;">
        Powered by Streamlit · 数据来源 Bilibili API
    </p>
</div>
""", unsafe_allow_html=True)
